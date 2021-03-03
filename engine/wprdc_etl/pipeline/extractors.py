import csv
import datetime
import io
from collections import OrderedDict
from engine.wprdc_etl.pipeline.exceptions import IsHeaderException
from xlrd import open_workbook, xldate_as_tuple, XL_CELL_DATE
from openpyxl import load_workbook

class Extractor(object):
    def __init__(self, connection):
        self.connection = connection
        self.checksum = None
        self.data = []

    def process_connection(self):
        '''Do any additional processing on a connection to prepare it for reading
        '''
        raise NotImplementedError

    def handle_line(self, line):
        '''Handle each line or row in a particular file.
        '''
        raise NotImplementedError

    def set_headers(self):
        '''Sets the column names or headers. Required for serialization
        '''
        raise NotImplementedError

class FileExtractor(Extractor):
    '''Extractor subclass for extracting entire files
    '''

    def __init__(self, connection, *args, **kwargs):
        super(FileExtractor, self).__init__(connection)

    def handle_line(self, line):
        '''Sets up the element from the iterable to be handled by the loader's load_line function.
        '''
        return line # Just return the file.

    def process_connection(self):
        '''This function gets the file returned from the Connector.connect()
        function and puts it in an iterator.
        '''
        reader = iter([self.connection])
        # Eventually update the Connector to support a list of files.
        # resource_names_list might have to be specified in the job_dict
        # and an AllOrSelectedFilesInADirectoryHTTPExtractor would be
        # needed to complete this kind of vectorized pipeline.
        return reader

class CompressedFileExtractor(FileExtractor):
    '''Extractor subclass for extracting files from a compressed file.
    '''

    def __init__(self, connection, *args, **kwargs):
        super(FileExtractor, self).__init__(connection)
        self.compressed_file_to_extract = kwargs.get('compressed_file_to_extract', None)

    def handle_line(self, line):
        '''Sets up the element from the iterable to be handled by the loader's load_line function.
        '''
        return line # Just return the file.

    def _uncompress(self):
        # Uses self.connection and self.compressed_file_to_extract.
        # Keep the file around in some form (maybe a temp file or memory).
        # Eventually the loader saves it to job.destination_file.
        from zipfile import ZipFile
        with ZipFile(self.connection, 'r') as zip_ref:
            with zip_ref.open(self.compressed_file_to_extract, 'r') as desired_file_ref:
                desired_file = io.BytesIO(desired_file_ref.read())
                return desired_file

    def process_connection(self):
        '''This function gets the file returned from the Connector.connect()
        function, uncompresses it to obtain the desired file, and puts the
        desired file in an iterator.
        '''
        if self.compressed_file_to_extract is not None:
            desired_file = self._uncompress()
            reader = iter([desired_file])
        else:
            reader = iter([])
        # Eventually update the Connector to support a list of files.
        # resource_names_list might have to be specified in the job_dict
        # and an AllOrSelectedFilesInADirectoryHTTPExtractor would be
        # needed to complete this kind of vectorized pipeline.
        return reader

class TableExtractor(Extractor):
    '''Abstract Extractor subclass for extracting data in a tabular format
    '''

    def __init__(self, connection, *args, **kwargs):
        super(TableExtractor, self).__init__(connection)
        self.headers = kwargs.get('headers', None)
        self.delimiter = kwargs.get('delimiter', ',')
        self.firstline_headers = kwargs.get('firstline_headers', True)

    def set_headers(self, headers=None):
        '''Sets headers from file or passed headers

        This method sets two attributes on the class: the headers
        attribute and the schema_headers attribute. schema_headers
        must align with the schema attributes for the pipeline.

        If no headers are passed, then we check the first line of the file.
        The headers attribute is set from the first line, and schema
        headers are by default created by lowercasing and replacing
        spaces with underscores. Custom header -> schema header mappings can
        be created by subclassing the CSVExtractor and overwriting the
        create_schema_headers method.

        Keyword Arguments:
            headers: Optional headers that can be passed to be used
                as the headers and schema headers

        Raises:
            RuntimeError: if self.headers is not passed and the
            firstline_headers kwarg is not set.
        '''
        if headers:
            self.headers = headers
            self.schema_headers = self.headers
            return
        elif self.firstline_headers:
            reader = self.process_connection()
            self.headers = next(reader)
            self.schema_headers = self.create_schema_headers(self.headers)
        else:
            raise RuntimeError('No headers were passed or detected.')

    def create_schema_headers(self, headers):
        '''Maps headers to schema headers

        Arguments:
            headers: A list of headers

        Returns:
            a list of formatted schema headers. By default,
                the passed headers are lowercased and have their
                spaces replaced with underscores
        '''
        return [
            i.lower().rstrip().replace(' ', '_').replace('-', '_') for i in
            headers
            ]

    def handle_line(self, line):
        '''Replace empty strings with None types.
        '''
        if line == self.headers:
            raise IsHeaderException
        return OrderedDict(zip(self.schema_headers, [i if i != '' else None for i in line]))


class CSVExtractor(TableExtractor):
    def __init__(self, connection, *args, **kwargs):
        '''TableExtractor subclass for CSV or character-delimited files
        '''
        super(CSVExtractor, self).__init__(connection, *args, **kwargs)
        self.delimiter = kwargs.get('delimiter', ',')
        self.set_headers()

    def process_connection(self):
        reader = csv.reader(self.connection, delimiter=self.delimiter)
        return reader

class ExcelExtractor(TableExtractor):
    '''TableExtractor subclass for newer Microsoft Excel spreadsheet files (XLSX)
    '''
    # XLSX files are compressed XML files.
    def __init__(self, connection, *args, **kwargs):
        super(ExcelExtractor, self).__init__(connection, *args, **kwargs)
        self.firstline_headers = kwargs.get('firstline_headers', True)
        self.sheet_index = kwargs.get('sheet_index', 0)
        self.rows_to_skip = kwargs.get('rows_to_skip', 0)
        self.datemode = None
        self.set_headers()

    def set_headers(self, headers=None):
        '''Sets headers from file or passed headers

        This method sets two attributes on the class: the headers
        attribute and the schema_headers attribute. schema_headers
        must align with the schema attributes for the pipeline.

        If no headers are passed, then we check the first line of the file.
        The headers attribute is set from the first line, and schema
        headers are by default created by lowercasing and replacing
        spaces with underscores. Custom header -> schema header mappings can
        be created by subclassing the CSVExtractor and overwriting the
        create_schema_headers method.

        Keyword Arguments:
            headers: Optional headers that can be passed to be used
                as the headers and schema headers

        Raises:
            RuntimeError: if self.headers is not passed and the
            firstline_headers kwarg is not set.
        '''
        if headers:
            self.headers = headers
            self.schema_headers = self.headers
            return
        elif self.firstline_headers:
            self.connection.seek(0) # This line doesn't seem to do anything for an Excel file.
            workbook = load_workbook(self.connection, read_only=True, data_only=True)
            # openpyxl's load_workbook function prefers to be sent a filename
            # or a file-like object open in binary mode e.g., zipfile.ZipFile.
            sheet = workbook.worksheets[self.sheet_index]
            self.headers = [cell.value for cell in sheet[1 + self.rows_to_skip]] # The first row in an Excel file is row 1.
            self.schema_headers = self.create_schema_headers(self.headers)
        else:
            raise RuntimeError('No headers were passed or detected.')

    def process_connection(self):
        data = []
        self.connection.seek(0) # This line doesn't seem to do anything for an Excel file.
        workbook = load_workbook(self.connection, read_only=True, data_only=True)
        sheet = workbook.worksheets[self.sheet_index]
        rows = sheet.rows
        data = []
        for k, row in enumerate(rows):
            if k < self.rows_to_skip:
                continue
            values = []
            for cell in row:
                if cell.data_type == 's':
                    values.append(cell.value.strip())
                else:
                    values.append(cell.value)
            data.append(values)
            if len(data) % 1000 == 0: print(len(data))

        return iter(data)

class OldExcelExtractor(TableExtractor):
    '''TableExtractor subclass for Microsoft Excel spreadsheet files (XLS)
    '''
    # Note that this supposedly used to support XLSX files (I think that
    # xlrd used to support XLSX files but no longer does).
    def __init__(self, connection, *args, **kwargs):
        super(OldExcelExtractor, self).__init__(connection, *args, **kwargs)
        self.firstline_headers = kwargs.get('firstline_headers', True)
        self.sheet_index = kwargs.get('sheet_index', 0)
        self.datemode = None
        self.set_headers()

    def process_connection(self):
        data = []
        self.connection.seek(0)
        contents = self.connection.read()
        workbook = open_workbook(file_contents=contents)
        sheet = workbook.sheet_by_index(self.sheet_index)
        self.datemode = workbook.datemode
        for i in range(sheet.nrows):
            data.append(self._read_line(sheet, i))

        return iter(data)

    def _read_line(self, sheet, row):
        '''Helper function to read line from Excel files and handle representations of
            different data types
        '''
        line = []
        for col in range(sheet.ncols):
            cell = sheet.cell(row, col)
            if cell.ctype == XL_CELL_DATE:
                xldate_tuple = xldate_as_tuple(cell.value, self.datemode)
                if xldate_tuple[0]:
                    date = datetime.datetime(*xldate_as_tuple(cell.value, self.datemode))
                    dt = date.strftime('%m/%d/%Y')  # todo: return datetime and handle the formatting elsewhere
                else:
                    time = datetime.time(*xldate_tuple[3:])
                    dt = time.strftime('%H:%M:%S')
                line.append(dt)
            else:
                line.append(cell.value)
        return line
